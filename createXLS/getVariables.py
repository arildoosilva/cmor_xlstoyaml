# -*- coding: utf-8 -*-
import re
import sys
import formatValues
from openpyxl import Workbook
from openpyxl import load_workbook

# vars
inpath = None
outpath = None
xlstable = None
wb = None
activetable = None
ignoredvars = ["eta", "depth", "sigma", "sigma_bnds", "zlev", "zlev_bnds", 
            "depth_c", "a", "b", "nsigma", "z1", "z2", "href", "k_c"]
            # ignores vertical coordinate formula_terms


def makeXLS(variable_entry, var_name):
    """ Creates a sheet for the variable it receives
    """
    variable_holder = []  # holds the list after the regex is applied
    variable = ""  # current variable
    variable_entry = variable_entry.split('\n')
    variable_entry = [x for x in variable_entry if not x.startswith(('!', ':'))]
    row = 2  # start on the second row of the sheet
    global wb
    global inpath
    global outpath

    for line in variable_entry:
        line = re.sub(' +', ' ', line)
        linekey = (line.partition(' ')[0]).replace(':', ' - ')
        linevalue = line.partition(' ')[2]
        line = linekey + linevalue
        variable_holder.append(re.sub(' +', ' ', line))
    variable_holder = filter(None, variable_holder)

    print("Adding variable: " + var_name)
    variable_sheet = wb.create_sheet(title=var_name)

    variable_sheet.cell(row=row-1, column=1).value = "table"
    variable_sheet.cell(row=row-1, column=2).value = '"{}"'.format(activetable)

    for line in variable_holder:
        line = line.replace(" - ", " = ", 1)
        try:
            key, value = line.split(" = ")
        except ValueError, e:
            print("\t" + line)
            print("\tValueError: check if the last processed string has a ' = '.")
            sys.exit(2)

        # Formats the key and value for proper yaml writing later
        newkey, newvalue = formatValues.format(key, value)

        variable_sheet.cell(row=row, column=1).value = newkey
        variable_sheet.cell(row=row, column=2).value = newvalue

        if newkey == "units":
            row += 1
            variable_sheet.cell(row=row, column=1).value = "unformatted units"
            variable_sheet.cell(row=row, column=2).value = newvalue

        row += 1

    bmgcs_vars_file = open(inpath + "/Variables_Used.txt")
    bmgcs_vars = bmgcs_vars_file.readlines()

    for line in bmgcs_vars:
        if var_name == line.split(", ")[0]:
            variable_sheet.cell(row=row, column=1).value = "name in BMGCS"
            variable_sheet.cell(row=row, column=2).value = '"{}"'.format(line.split(", ")[1])
            variable_sheet.cell(row=row+1, column=1).value = "units in BMGCS"
            variable_sheet.cell(row=row+1, column=2).value = '"{}"'.format(line.split(", ")[2])
            break

    # Writes some secondary rows
    variable_sheet.cell(row=row+2, column=1).value = "priority"
    variable_sheet.cell(row=row+2, column=2).value = 1
    variable_sheet.cell(row=row+3, column=1).value = "questions and notes"
    variable_sheet.cell(row=row+3, column=2).value = ""
    variable_sheet.cell(row=row+4, column=1).value = "positive"
    variable_sheet.cell(row=row+4, column=2).value = ""
    variable_sheet.cell(row=row+5, column=1).value = "frequency"
    variable_sheet.cell(row=row+5, column=2).value = ""
    variable_sheet.cell(row=row+6, column=1).value = "flag_values"
    variable_sheet.cell(row=row+6, column=2).value = ""
    variable_sheet.cell(row=row+7, column=1).value = "flag_meanings"
    variable_sheet.cell(row=row+7, column=2).value = ""

    wb.save(filename=xlstable)


def getTheSentences(infile):
    """ Uses regex to find the variables in every file, then
    send each variable to makeXLS
    """
    global ignoredvars
    with open(infile) as fp:
        for variable_entry in re.findall('variable_entry(.*?)variable_entry', fp.read(), re.S):
            for each in variable_entry.split("\n"):
                var_name = each.split()[1]
                break
            if var_name not in ignoredvars:
                makeXLS(variable_entry, var_name)


def readTables(indir, outdir, tables):
    """ Reads the table list and send each file from the list to the
    getTheSentences def
    """
    global inpath
    inpath = indir
    global outpath
    outpath = outdir
    for item in tables:
        global xlstable
        global wb
        global activetable
        activetable = item
        xlstable = outdir + "/" + item[6:] + ".xlsx"
        wb = load_workbook(filename=xlstable)
        print("\n" + item + "\n")  # current table
        filename = inpath + '/' + item
        getTheSentences(filename)
