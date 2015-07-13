# -*- coding: utf-8 -*-
import sys
import getopt
import os
import fileCheck
import getVariables
import createYaml
from openpyxl import Workbook

# vars
inpath = None
outpath = None
table_prefix = 'CMIP5_'  # change this when using CMIP6 tables
tables = []
wb = Workbook()


# defs
def getArgs(argv):
    """ Gets the arguments passed in terminal
    """
    global inpath
    global outpath

    help_text = ('This script will generate XLS files from CMOR tables\n'
            'Modules used: openpyxl, pyyaml, sys, getopt, os')

    # prints the help text when -h or --help is used
    try:
        opts, args = getopt.getopt(argv, 'h', ['help'])
    except get.GetoptError:
        print(help_text)
        sys.exit(2)

    for opt, arg in opts:
        if opt in ('-h', '--help'):
            print(help_text)
            sys.exit()

    if len(sys.argv) < 3:
        print('Missing parameters\n')
        sys.exit(2)
    else:
        inpath = sys.argv[1]
        outpath = sys.argv[2]


def readIn(inpath):
    """ Gets the tables from the IN directory
    """
    global tables
    global table_prefix
    for (dirpath, dirnames, filenames) in os.walk(inpath):
        for f in filenames:
            if (table_prefix in f and "grids" not in f):
                print("Adding file " + f + " to list")
                tables.append(f)


def writeDescription(inpath, outpath, tables):
    """ Writes the description header on each xls file
    """
    global wb
    row = 1  # start on the first row of the sheet
    description_sheet = wb.active
    description_sheet.title = "description"
    description_file = open((inpath + "/Description.txt"), "r")
    print("\nReading the description file\n")
    for line in description_file:
        line = line.replace("\n", "")
        try:
            key, value = line.split(" = ")
        except ValueError, e:
            print("\t" + line)
            print("\tValueError: check if the last processed string has a ' = '.")
            sys.exit(2)
        # print("Adding line: " + line)  # DEBUG
        description_sheet.cell(row=row, column=1).value = key
        description_sheet.cell(row=row, column=2).value = value
        row += 1
    for item in tables:
        filename = outpath + "/" + item[6:] + ".xlsx"
        wb.save(filename = filename)


# main
if __name__ == "__main__":
    getArgs(sys.argv[0:])
    try:
        readIn(inpath)
        writeDescription(inpath, outpath, tables)
        getVariables.readTables(inpath, outpath, tables)
    except:
        e = sys.exc_info()[0]
        print(e)
        sys.exit(2)
    answer = raw_input("Do you want to generate the tables now? [yes/no] ")
    if "yes" in answer:
        createYaml.main(outpath)
        createYaml.clearOut()  # deletes all xlsx files, leaving only yaml files
